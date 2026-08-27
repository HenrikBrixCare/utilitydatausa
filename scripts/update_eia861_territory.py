#!/usr/bin/env python3
"""Generate a compact county -> candidate electric utility index from official EIA-861 bulk data.

This dataset is intentionally county-level context only. EIA describes Service Territory as
counties/states where a utility has equipment to distribute electricity. It is not proof that a
specific utility serves a specific street address.
"""

from __future__ import annotations

import io
import json
import os
import re
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

YEAR = int(os.environ.get("EIA861_YEAR", "2024"))
SOURCE_URL = f"https://www.eia.gov/electricity/data/eia861/zip/f861{YEAR}.zip"
OUTPUT = Path(os.environ.get("EIA861_OUTPUT", f"data/eia861_service_territory_{YEAR}.json"))


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm(value).lower())


def download() -> bytes:
    request = urllib.request.Request(
        SOURCE_URL,
        headers={"User-Agent": "UtilityDataUSA-EIA861-generator/0.1"},
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def find_workbook(archive: zipfile.ZipFile) -> str:
    wanted = f"service_territory_{YEAR}.xlsx".lower()
    for name in archive.namelist():
        if Path(name).name.lower() == wanted:
            return name
    candidates = [name for name in archive.namelist() if "service_territory" in name.lower() and name.lower().endswith(".xlsx")]
    if not candidates:
        raise RuntimeError("Service Territory workbook not found in EIA-861 ZIP")
    return candidates[0]


def locate_header(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    aliases = {
        "utility_id": {"utilitynumber", "utilityid", "utilityidnumber", "utilitynumberid"},
        "utility_name": {"utilityname", "entityname"},
        "state": {"state", "stateabbreviation", "stateabbr"},
        "county": {"county", "countyname"},
    }
    for row_index, row in enumerate(rows[:30]):
        normalized = [header_key(cell) for cell in row]
        mapping: dict[str, int] = {}
        for logical, choices in aliases.items():
            for index, key in enumerate(normalized):
                if key in choices:
                    mapping[logical] = index
                    break
        if {"utility_id", "utility_name", "state", "county"}.issubset(mapping):
            return row_index, mapping
    raise RuntimeError(f"Could not identify expected EIA-861 columns. First rows: {rows[:5]!r}")


def main() -> None:
    raw_zip = download()
    with zipfile.ZipFile(io.BytesIO(raw_zip)) as archive:
        workbook_name = find_workbook(archive)
        workbook_bytes = archive.read(workbook_name)

    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header_row, columns = locate_header(rows)

    grouped: dict[str, dict[str, object]] = defaultdict(lambda: {"utilities": []})
    seen: set[tuple[str, str, str]] = set()

    for row in rows[header_row + 1 :]:
        state = norm(row[columns["state"]]).upper()
        county = norm(row[columns["county"]])
        utility_id = norm(row[columns["utility_id"]])
        utility_name = norm(row[columns["utility_name"]])
        if not state or not county or not utility_name:
            continue

        # EIA files may contain notes/totals; state abbreviations give us a simple guardrail.
        if not re.fullmatch(r"[A-Z]{2}", state):
            continue

        key = f"{state}|{county.upper()}"
        dedupe = (key, utility_id, utility_name)
        if dedupe in seen:
            continue
        seen.add(dedupe)

        entry = grouped[key]
        entry["state"] = state
        entry["county"] = county
        entry["utilities"].append({"utility_id_eia": utility_id or None, "utility_name": utility_name})

    counties = []
    for key in sorted(grouped):
        entry = grouped[key]
        entry["utilities"] = sorted(entry["utilities"], key=lambda item: (item["utility_name"].lower(), item["utility_id_eia"] or ""))
        counties.append(entry)

    payload = {
        "source": "U.S. Energy Information Administration Form EIA-861 Service Territory",
        "source_url": SOURCE_URL,
        "year": YEAR,
        "scope": "County-level candidate electric distribution utilities",
        "limitation": "EIA-861 Service Territory lists counties/states where a utility has equipment to distribute electricity. More than one utility can appear in a county. This is not proof that a utility serves a particular street address.",
        "county_count": len(counties),
        "records": counties,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Generated {OUTPUT} with {len(counties)} county/state entries from {workbook_name}")
    sample = counties[:3]
    print("Sample:", json.dumps(sample, ensure_ascii=False)[:1500])


if __name__ == "__main__":
    main()
